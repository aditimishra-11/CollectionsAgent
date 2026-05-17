"""Build a single xlsx deliverable from the five result CSVs in the repo.

Output: 'Aditi Mishra - Eval Results.xlsx' at repo root.

Tabs:
  1. Index — overview + how to read
  2. v1_baseline — original v1 runner output (4 axes)
  3. v1_backfilled — v1 transcripts re-graded with v2's judge (9 axes)
  4. v2_synthetic — v2 runner output on 42 scenarios
  5. v2_real_calls — runner_live.py output on 24 real recorded calls
  6. v1_vs_v2_shared — apples-to-apples comparison on the 15 shared scenarios
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
OUT = ROOT / "Aditi Mishra - Eval Results.xlsx"

V1 = ROOT / "collections-voicebot" / "eval" / "results_v1.csv"
V1_BF = ROOT / "collections-voicebot" / "eval" / "results_v1_backfilled.csv"
V2 = ROOT / "collections-voicebot-v2" / "eval" / "results_v2.csv"
LIVE = ROOT / "collections-voicebot-v2" / "eval" / "results_live.csv"
CMP = ROOT / "collections-voicebot-v2" / "eval" / "comparison_v1_v2.csv"

# ---------- style helpers ----------

HEADER_FILL = PatternFill("solid", fgColor="0E2E5C")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SECTION_FONT = Font(bold=True, size=12, color="0E2E5C")
BORDER_THIN = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def style_header_row(ws, row_idx: int, n_cols: int) -> None:
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER_THIN


def autosize_columns(ws, max_width: int = 60) -> None:
    for col_cells in ws.columns:
        first = next((c for c in col_cells if c.value is not None), None)
        if first is None:
            continue
        col_letter = get_column_letter(first.column)
        max_len = 0
        for c in col_cells:
            if c.value is None:
                continue
            val_str = str(c.value)
            longest_line = max((len(line) for line in val_str.split("\n")), default=0)
            max_len = max(max_len, longest_line)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def write_dataframe(ws, df: pd.DataFrame, freeze_header: bool = True) -> None:
    """Write df starting at row 1, with styled header."""
    # Header
    for c, col in enumerate(df.columns, 1):
        ws.cell(row=1, column=c, value=str(col))
    style_header_row(ws, 1, len(df.columns))
    # Data
    for r, row in enumerate(df.itertuples(index=False), 2):
        for c, val in enumerate(row, 1):
            if pd.isna(val):
                cell_val = ""
            elif isinstance(val, (int, float)):
                cell_val = val
            else:
                cell_val = str(val)
            cell = ws.cell(row=r, column=c, value=cell_val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.font = Font(size=10)
    autosize_columns(ws)
    if freeze_header:
        ws.freeze_panes = "A2"


# ---------- index tab ----------

INDEX_ROWS = [
    ("title", "Mumbai Bank Collections Voicebot — Eval Results", None),
    ("subtitle", "Aditi Mishra · GreyLabs AI PM Assignment · 17 May 2026", None),
    ("blank", "", ""),
    ("section", "What's in this workbook", ""),
    ("text",
     "Five tabs of evaluation data covering v1 (baseline) and v2 (production architecture) of the bot, "
     "plus a side-by-side comparison. The doc 'Aditi Mishra - Assignment Writeup' references these tabs "
     "for the detailed per-scenario evidence behind the headline numbers.", ""),
    ("blank", "", ""),
    ("section", "Tabs and how to read them", ""),
    ("blank", "", ""),
    ("table_head", "Tab", "What it contains and how to read it"),

    ("table_row", "v1_baseline",
     "v1 = the literal starter prompt from the assignment brief, unmodified. Run through the original 4-axis eval runner "
     "(at collections-voicebot/eval/runner.py). 15 scripted scenarios. Columns: scenario_id, persona_id, bucket, difficulty, "
     "expected_outcome, actual_outcome, outcome_match, compliance_pass, compliance_failures, tone_pass, tone_failures, "
     "escalation_correct, turns, transcript_path, notes. KEY READ: compliance_pass is the headline failure — only 13% of v1 "
     "calls (2 of 15) passed without a policy violation."),

    ("table_row", "v1_backfilled",
     "Same 15 v1 transcripts re-graded with v2's GPT-4o judge / extractor to fill in axes the v1 runner did not capture "
     "(hallucination, PTP specificity, containment, Likert empathy / sentiment / context retention). Source: backfill script at "
     "collections-voicebot-v2/eval/backfill_v1.py. Use this tab when comparing v1 ↔ v2 on the experience-layer axes that the "
     "original v1 runner did not record. NOT backfillable: cost and latency (no per-turn token / timestamp data was stored)."),

    ("table_row", "v2_synthetic",
     "v2 production architecture, run on 42 scripted scenarios. Source: collections-voicebot-v2/eval/runner.py. "
     "9 axes plus efficiency metrics. KEY COLUMNS: priority (P0 / P1 / P2 — weighted 3× / 2× / 1× respectively), "
     "compliance_pass, outcome_match, tone_pass, transfer_correct, hallucination_pass, slot_capture_rate, slots_captured, "
     "slots_missed, empathy_score / sentiment_trajectory / context_retention (Likert 0-5; >=3 counts as pass), "
     "llm_p50_ms / llm_p95_ms / estimated_inr_per_call, full_pass (all 6 binary axes AND-ed). "
     "KEY READ: P0 compliance is 100% on this set; overall full_pass is 52%."),

    ("table_row", "v2_real_calls",
     "v2 graded on 24 real recorded JSONL transcripts in logs/. Source: collections-voicebot-v2/eval/runner_live.py. "
     "Auto-annotation: ground truth (expected_outcome, should_transfer) inferred at call end by an isolated LLM-judge prompt "
     "and persisted to disk. Hand-written annotations in eval/annotations_live.yaml take precedence where they exist. "
     "ADDITIONAL COLUMNS NOT IN SYNTHETIC: contract_consistency_pass, closure_coherence_pass (real-call-only axes), "
     "bot_must_pass_count / bot_must_total, bot_must_not_pass_count / bot_must_not_total. "
     "KEY READ: outcome_match 42% on this set is lower than synthetic (76%) because the LLM-inferred ground truth is stricter; "
     "compliance_pass 92% and tone_pass 96% remain close to synthetic numbers."),

    ("table_row", "v1_vs_v2_shared",
     "Apples-to-apples per-scenario comparison on the 15 shared scenarios (the original v1 set). Side-by-side columns "
     "for each axis (e.g., compliance_pass_v1 next to compliance_pass_v2). Use this tab to see exactly which scenarios "
     "behaved differently between v1 and v2. KEY READ: compliance_pass moves from 13% v1 to 100% v2 on the shared 15."),

    ("blank", "", ""),
    ("section", "Pass / fail thresholds and definitions", ""),
    ("blank", "", ""),
    ("table_head", "Concept", "Definition"),
    ("table_row", "full_pass",
     "Strict ALL-axes-AND-ed gate. v1: 4 axes (outcome_match + compliance_pass + tone_pass + escalation_correct). "
     "v2 synthetic: 6 axes (the 4 above + hallucination_pass + slot_capture_rate==1.0)."),
    ("table_row", "compliance_pass",
     "Binary. True if the validator's 17 rules ALL passed for the call. compliance_failures lists which rules fired."),
    ("table_row", "tone_pass",
     "Binary. LLM judge on GPT-4o decides whether the tone matched what the segment requires (concierge for Apex, firm for frequent-late, etc.)."),
    ("table_row", "outcome_match",
     "Binary. actual_outcome == expected_outcome. v2 may legitimately diverge from expected because segment-policy thresholds push back on too-far PTPs."),
    ("table_row", "transfer_correct / escalation_correct",
     "Binary. Did the bot escalate when it should have (medical / job-loss / abuse / dispute) and not when it shouldn't have?"),
    ("table_row", "slot_capture_rate",
     "0.0 to 1.0. For PTP calls: fraction of required slots (date, mode) captured. PTP specificity = 1.0 means both date and mode were captured."),
    ("table_row", "Likert axes (empathy / sentiment / context_retention)",
     "0 to 5 from the GPT-4o judge. >=3 counts as pass when used as a binary axis. If the judge call fails the runner defaults to 3 (a known bias the writeup discloses)."),
    ("table_row", "priority (P0 / P1 / P2)",
     "P0 = zero-tolerance scenarios (single violation triggers RBI Fair Practices Code review; 16 scenarios). "
     "P1 = high-priority (26 scenarios). P2 = standard. Weighted 3× / 2× / 1× in any priority-aggregate metric."),

    ("blank", "", ""),
    ("section", "Source files", ""),
    ("table_head", "Tab", "Source CSV path"),
    ("table_row", "v1_baseline", "collections-voicebot/eval/results_v1.csv"),
    ("table_row", "v1_backfilled", "collections-voicebot/eval/results_v1_backfilled.csv"),
    ("table_row", "v2_synthetic", "collections-voicebot-v2/eval/results_v2.csv"),
    ("table_row", "v2_real_calls", "collections-voicebot-v2/eval/results_live.csv"),
    ("table_row", "v1_vs_v2_shared", "collections-voicebot-v2/eval/comparison_v1_v2.csv"),
]


def write_index(ws) -> None:
    row = 1
    for kind, a, b in INDEX_ROWS:
        if kind == "title":
            cell = ws.cell(row=row, column=1, value=a)
            cell.font = Font(bold=True, size=18, color="0E2E5C")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
        elif kind == "subtitle":
            cell = ws.cell(row=row, column=1, value=a)
            cell.font = Font(italic=True, size=11, color="555555")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
        elif kind == "section":
            cell = ws.cell(row=row, column=1, value=a)
            cell.font = SECTION_FONT
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
        elif kind == "text":
            cell = ws.cell(row=row, column=1, value=a)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.font = Font(size=10)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            ws.row_dimensions[row].height = 60
            row += 1
        elif kind == "table_head":
            c1 = ws.cell(row=row, column=1, value=a)
            c2 = ws.cell(row=row, column=2, value=b)
            c1.fill = HEADER_FILL
            c1.font = HEADER_FONT
            c1.alignment = Alignment(vertical="center", horizontal="left")
            c2.fill = HEADER_FILL
            c2.font = HEADER_FONT
            c2.alignment = Alignment(vertical="center", horizontal="left")
            c1.border = BORDER_THIN
            c2.border = BORDER_THIN
            row += 1
        elif kind == "table_row":
            c1 = ws.cell(row=row, column=1, value=a)
            c2 = ws.cell(row=row, column=2, value=b)
            c1.font = Font(bold=True, size=10)
            c1.alignment = Alignment(vertical="top", wrap_text=True)
            c2.font = Font(size=10)
            c2.alignment = Alignment(vertical="top", wrap_text=True)
            c1.border = BORDER_THIN
            c2.border = BORDER_THIN
            # Height by content length
            text_len = len(b) if b else 0
            ws.row_dimensions[row].height = max(20, min(text_len // 4, 200))
            row += 1
        elif kind == "blank":
            row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110


# ---------- main ----------

def main() -> None:
    wb = Workbook()
    # Index tab
    idx = wb.active
    idx.title = "Index"
    write_index(idx)

    # Data tabs
    tabs = [
        ("v1_baseline", V1),
        ("v1_backfilled", V1_BF),
        ("v2_synthetic", V2),
        ("v2_real_calls", LIVE),
        ("v1_vs_v2_shared", CMP),
    ]
    for sheet_name, path in tabs:
        if not path.exists():
            print(f"WARN: {path} not found, skipping {sheet_name}")
            continue
        df = pd.read_csv(path)
        ws = wb.create_sheet(title=sheet_name)
        write_dataframe(ws, df)
        print(f"  {sheet_name}: {len(df)} rows, {len(df.columns)} cols")

    wb.save(OUT)
    print(f"\nWrote {OUT} ({OUT.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
